import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

driver = webdriver.Chrome(executable_path='C:/Users/amuri/AppData/Local/Microsoft/WindowsApps/PythonSoftwareFoundation.Python.3.9_qbz5n2kfra8p0/site-packages/chromedriver.exe')
driver.implicitly_wait(1)


def get_comp_type(comp_pn):
    url ='https://www.digikey.com/'
    driver.get(url)
    print(driver.title)
    wait = WebDriverWait(driver, timeout=1)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#main-layout-content .header__searchinput")))
    elem = driver.find_element_by_css_selector("#main-layout-content .header__searchinput")
    elem.click()
    elem.send_keys(comp_pn)
    elem.send_keys(Keys.RETURN)
    time.sleep(1)
    comp_type = driver.find_element_by_xpath("/html/body/div[2]/main/div/div[2]/div[1]/div[3]/div/table/tbody/tr[18]/td[2]/div/div").text
    if comp_type == '2':
        value_2 = driver.find_element_by_xpath("/html/body/div[2]/main/div/div[1]/div[1]/div[3]/div/table/tbody/tr[19]/td[2]/div/div").text
        print(value_2)
        return value_2
    else:
        print(comp_type)
        return comp_type


wb = load_workbook('desco.xlsx')
sheet = wb['Resistors']

total_company_parts = 363
for i in range (195, total_company_parts+1):
    mfr_pn = sheet.cell(row=i, column=5).value
    # Try block used in order to prevent complete script from crashing
    # if some component not found in Digikey
    try:
        sheet.cell(row=i, column=8).value = get_comp_type(mfr_pn)
        wb.save("desco.xlsx")
    except:
        continue

